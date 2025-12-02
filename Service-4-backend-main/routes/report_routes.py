from flask import Blueprint, request, jsonify, send_file
from flask_login import login_required, current_user
from models import Order, OrderItem, Customer, Product, Invoice, InvoiceItem, get_db
from bson import ObjectId
from datetime import datetime, timedelta
from collections import defaultdict
from io import BytesIO
try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False
try:
    from reportlab.lib.pagesizes import letter, A4
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet
    from reportlab.lib import colors
    REPORTLAB_AVAILABLE = True
except ImportError:
    REPORTLAB_AVAILABLE = False

report_bp = Blueprint('report', __name__)

@report_bp.route('/api/sales-summary', methods=['GET'])
@login_required
def sales_summary():
    """Get sales summary with revenue, orders, customers"""
    try:
        database = get_db()
        if database is None:
            return jsonify({'success': False, 'error': 'Database not initialized'}), 500
        
        # Get user_id and filter by it
        user_id = current_user.id
        user_id_obj = ObjectId(user_id) if isinstance(user_id, str) and ObjectId.is_valid(user_id) else user_id
        
        # Get date range (default: last 30 days)
        days = request.args.get('days', 30, type=int)
        start_date = datetime.now() - timedelta(days=days)
        
        # Total revenue from invoices (not orders, since we're using invoices)
        revenue_pipeline = [
            {'$match': {
                'user_id': user_id_obj,
                'created_at': {'$gte': start_date}
            }},
            {'$group': {'_id': None, 'total': {'$sum': '$total_amount'}}}
        ]
        revenue_result = list(database['invoices'].aggregate(revenue_pipeline))
        total_revenue = revenue_result[0].get('total', 0) if revenue_result else 0.0
        
        # Total invoices
        total_orders = database['invoices'].count_documents({
            'user_id': user_id_obj,
            'created_at': {'$gte': start_date}
        })
        
        # Total customers for this user
        total_customers = database['customers'].count_documents({'user_id': user_id_obj})
        
        # Active customers (placed invoices in period)
        active_customers_pipeline = [
            {'$match': {
                'user_id': user_id_obj,
                'created_at': {'$gte': start_date}
            }},
            {'$group': {'_id': '$customer_id'}},
            {'$count': 'active_customers'}
        ]
        active_result = list(database['invoices'].aggregate(active_customers_pipeline))
        active_customers = active_result[0].get('active_customers', 0) if active_result else 0
        
        # Average order value
        avg_order_value = total_revenue / total_orders if total_orders > 0 else 0
        
        # Invoices by status
        orders_by_status_pipeline = [
            {'$match': {
                'user_id': user_id_obj,
                'created_at': {'$gte': start_date}
            }},
            {'$group': {
                '_id': '$status',
                'count': {'$sum': 1},
                'revenue': {'$sum': '$total_amount'}
            }}
        ]
        orders_by_status = list(database['invoices'].aggregate(orders_by_status_pipeline))
        
        status_breakdown = {
            item['_id']: {
                'count': item.get('count', 0),
                'revenue': float(item.get('revenue', 0))
            }
            for item in orders_by_status
        }
        
        return jsonify({
            'success': True,
            'summary': {
                'total_revenue': float(total_revenue),
                'total_orders': total_orders,
                'total_customers': total_customers,
                'active_customers': active_customers,
                'avg_order_value': float(avg_order_value),
                'status_breakdown': status_breakdown
            }
        })
    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'error': str(e)}), 500

@report_bp.route('/api/sales-trends', methods=['GET'])
@login_required
def sales_trends():
    """Get sales trends over time (daily, weekly, monthly)"""
    try:
        database = get_db()
        if database is None:
            return jsonify({'success': False, 'error': 'Database not initialized'}), 500
        
        user_id = current_user.id
        user_id_obj = ObjectId(user_id) if isinstance(user_id, str) and ObjectId.is_valid(user_id) else user_id
        
        period = request.args.get('period', 'daily')  # daily, weekly, monthly
        days = request.args.get('days', 30, type=int)
        start_date = datetime.now() - timedelta(days=days)
        
        base_match = {
            'user_id': user_id_obj,
            'created_at': {'$gte': start_date}
        }
        
        if period == 'daily':
            # Group by day
            trends_pipeline = [
                {'$match': base_match},
                {'$group': {
                    '_id': {'$dateToString': {'format': '%Y-%m-%d', 'date': '$created_at'}},
                    'orders': {'$sum': 1},
                    'revenue': {'$sum': '$total_amount'}
                }},
                {'$sort': {'_id': 1}}
            ]
            trends = list(database['invoices'].aggregate(trends_pipeline))
            
            data = [{
                'date': trend['_id'],
                'orders': trend.get('orders', 0),
                'revenue': float(trend.get('revenue', 0))
            } for trend in trends]
            
        elif period == 'weekly':
            # Group by week
            trends_pipeline = [
                {'$match': base_match},
                {'$group': {
                    '_id': {
                        'year': {'$year': '$created_at'},
                        'week': {'$week': '$created_at'}
                    },
                    'orders': {'$sum': 1},
                    'revenue': {'$sum': '$total_amount'}
                }},
                {'$sort': {'_id.year': 1, '_id.week': 1}}
            ]
            trends = list(database['invoices'].aggregate(trends_pipeline))
            
            data = [{
                'period': f"Week {trend['_id']['week']}, {trend['_id']['year']}",
                'orders': trend.get('orders', 0),
                'revenue': float(trend.get('revenue', 0))
            } for trend in trends]
            
        else:  # monthly
            # Group by month
            trends_pipeline = [
                {'$match': base_match},
                {'$group': {
                    '_id': {
                        'year': {'$year': '$created_at'},
                        'month': {'$month': '$created_at'}
                    },
                    'orders': {'$sum': 1},
                    'revenue': {'$sum': '$total_amount'}
                }},
                {'$sort': {'_id.year': 1, '_id.month': 1}}
            ]
            trends = list(database['invoices'].aggregate(trends_pipeline))
            
            data = [{
                'period': f"{trend['_id']['month']}/{trend['_id']['year']}",
                'orders': trend.get('orders', 0),
                'revenue': float(trend.get('revenue', 0))
            } for trend in trends]
        
        return jsonify({
            'success': True,
            'period': period,
            'data': data
        })
    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'error': str(e)}), 500

@report_bp.route('/api/top-customers', methods=['GET'])
@login_required
def top_customers():
    """Get top customers by revenue"""
    try:
        database = get_db()
        if database is None:
            return jsonify({'success': False, 'error': 'Database not initialized'}), 500
        
        user_id = current_user.id
        user_id_obj = ObjectId(user_id) if isinstance(user_id, str) and ObjectId.is_valid(user_id) else user_id
        
        limit = request.args.get('limit', 10, type=int)
        days = request.args.get('days', 30, type=int)
        start_date = datetime.now() - timedelta(days=days)
        
        top_customers_pipeline = [
            {'$match': {
                'user_id': user_id_obj,
                'created_at': {'$gte': start_date}
            }},
            {'$group': {
                '_id': '$customer_id',
                'order_count': {'$sum': 1},
                'total_spent': {'$sum': '$total_amount'}
            }},
            {'$sort': {'total_spent': -1}},
            {'$limit': limit},
            {'$lookup': {
                'from': 'customers',
                'localField': '_id',
                'foreignField': '_id',
                'as': 'customer'
            }},
            {'$unwind': {'path': '$customer', 'preserveNullAndEmptyArrays': True}}
        ]
        top_customers = list(database['invoices'].aggregate(top_customers_pipeline))
        
        customers_data = []
        for item in top_customers:
            customer = item.get('customer', {})
            customers_data.append({
                'id': str(item['_id']),
                'name': customer.get('name', 'Unknown'),
                'email': customer.get('email', ''),
                'order_count': item.get('order_count', 0),
                'total_spent': float(item.get('total_spent', 0))
            })
        
        return jsonify({
            'success': True,
            'customers': customers_data
        })
    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'error': str(e)}), 500

@report_bp.route('/api/top-products', methods=['GET'])
@login_required
def top_products():
    """Get top products by quantity sold"""
    try:
        database = get_db()
        if database is None:
            return jsonify({'success': False, 'error': 'Database not initialized'}), 500
        
        user_id = current_user.id
        user_id_obj = ObjectId(user_id) if isinstance(user_id, str) and ObjectId.is_valid(user_id) else user_id
        
        limit = request.args.get('limit', 10, type=int)
        days = request.args.get('days', 30, type=int)
        start_date = datetime.now() - timedelta(days=days)
        
        top_products_pipeline = [
            {'$match': {
                'user_id': user_id_obj,
                'created_at': {'$gte': start_date}
            }},
            {'$unwind': '$items'},
            {'$group': {
                '_id': '$items.product_id',
                'quantity_sold': {'$sum': '$items.quantity'},
                'revenue': {'$sum': '$items.total'}
            }},
            {'$sort': {'quantity_sold': -1}},
            {'$limit': limit},
            {'$lookup': {
                'from': 'products',
                'localField': '_id',
                'foreignField': '_id',
                'as': 'product'
            }},
            {'$unwind': {'path': '$product', 'preserveNullAndEmptyArrays': True}}
        ]
        top_products = list(database['invoices'].aggregate(top_products_pipeline))
        
        products_data = []
        for item in top_products:
            product = item.get('product', {})
            products_data.append({
                'id': str(item['_id']),
                'name': product.get('name', 'Unknown'),
                'sku': product.get('sku', ''),
                'quantity_sold': int(item.get('quantity_sold', 0)),
                'revenue': float(item.get('revenue', 0))
            })
        
        return jsonify({
            'success': True,
            'products': products_data
        })
    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'error': str(e)}), 500

@report_bp.route('/revenue-by-category', methods=['GET'])
@login_required
def revenue_by_category():
    """Get revenue breakdown by product category"""
    try:
        days = request.args.get('days', 30, type=int)
        start_date = datetime.now() - timedelta(days=days)
        
        category_revenue_pipeline = [
            {'$match': {'created_at': {'$gte': start_date}}},
            {'$unwind': '$items'},
            {'$lookup': {
                'from': 'products',
                'localField': 'items.product_id',
                'foreignField': '_id',
                'as': 'product'
            }},
            {'$unwind': {'path': '$product', 'preserveNullAndEmptyArrays': True}},
            {'$match': {
                '$or': [
                    {'product.category': {'$ne': None}},
                    {'product.category': {'$ne': ''}}
                ]
            }},
            {'$group': {
                '_id': '$product.category',
                'revenue': {'$sum': '$items.total'},
                'quantity': {'$sum': '$items.quantity'}
            }},
            {'$sort': {'revenue': -1}}
        ]
        category_revenue = list(db['orders'].aggregate(category_revenue_pipeline))
        
        data = [{
            'category': item['_id'] or 'Uncategorized',
            'revenue': float(item.get('revenue', 0)),
            'quantity': int(item.get('quantity', 0))
        } for item in category_revenue]
        
        return jsonify({
            'success': True,
            'data': data
        })
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500

@report_bp.route('/customer-growth', methods=['GET'])
@login_required
def customer_growth():
    """Get customer growth over time"""
    try:
        days = request.args.get('days', 90, type=int)
        start_date = datetime.now() - timedelta(days=days)
        
        growth_pipeline = [
            {'$match': {'created_at': {'$gte': start_date}}},
            {'$group': {
                '_id': {'$dateToString': {'format': '%Y-%m-%d', 'date': '$created_at'}},
                'new_customers': {'$sum': 1}
            }},
            {'$sort': {'_id': 1}}
        ]
        growth = list(db['customers'].aggregate(growth_pipeline))
        
        data = [{
            'date': item['_id'],
            'new_customers': item.get('new_customers', 0)
        } for item in growth]
        
        return jsonify({
            'success': True,
            'data': data
        })
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500

@report_bp.route('/api/download', methods=['GET'])
@login_required
def download_report():
    """Download reports as PDF or Excel"""
    try:
        format_type = request.args.get('format', 'excel')  # 'excel' or 'pdf'
        report_type = request.args.get('type', 'summary')  # 'summary', 'customers', 'products', 'trends'
        days = request.args.get('days', 30, type=int)
        start_date = datetime.now() - timedelta(days=days)
        
        # Get database instance and current user filter
        database = get_db()
        if database is None:
            return jsonify({'success': False, 'error': 'Database not initialized'}), 500
        
        # Filter by current user's business (user_id) for multi-tenant safety
        user_id = current_user.id
        user_id_obj = ObjectId(user_id) if isinstance(user_id, str) and ObjectId.is_valid(user_id) else user_id
        
        if format_type == 'excel' and OPENPYXL_AVAILABLE:
            wb = Workbook()
            
            if report_type == 'summary':
                ws = wb.active
                ws.title = "Sales Summary"
                
                # Get summary data
                revenue_pipeline = [
                    {'$match': {
                        'user_id': user_id_obj,
                        'created_at': {'$gte': start_date}
                    }},
                    {'$group': {'_id': None, 'total': {'$sum': '$total_amount'}}}
                ]
                revenue_result = list(database['invoices'].aggregate(revenue_pipeline))
                total_revenue = revenue_result[0].get('total', 0) if revenue_result else 0.0
                
                total_orders = database['invoices'].count_documents({
                    'user_id': user_id_obj,
                    'created_at': {'$gte': start_date}
                })
                total_customers = database['customers'].count_documents({'user_id': user_id_obj})
                
                active_customers_pipeline = [
                    {'$match': {
                        'user_id': user_id_obj,
                        'created_at': {'$gte': start_date}
                    }},
                    {'$group': {'_id': '$customer_id'}},
                    {'$count': 'active_customers'}
                ]
                active_result = list(database['invoices'].aggregate(active_customers_pipeline))
                active_customers = active_result[0].get('active_customers', 0) if active_result else 0
                
                # Write summary
                ws['A1'] = 'Sales Summary Report'
                ws['A1'].font = Font(bold=True, size=16)
                ws['A2'] = f'Period: Last {days} days'
                ws['A2'].font = Font(bold=True)
                
                ws['A4'] = 'Metric'
                ws['B4'] = 'Value'
                ws['A4'].font = Font(bold=True)
                ws['B4'].font = Font(bold=True)
                
                ws['A5'] = 'Total Revenue'
                ws['B5'] = float(total_revenue)
                ws['A6'] = 'Total Orders'
                ws['B6'] = total_orders
                ws['A7'] = 'Total Customers'
                ws['B7'] = total_customers
                ws['A8'] = 'Active Customers'
                ws['B8'] = active_customers
                
            elif report_type == 'customers':
                ws = wb.active
                ws.title = "Top Customers"
                
                limit = request.args.get('limit', 50, type=int)
                top_pipeline = [
                    {'$match': {
                        'user_id': user_id_obj,
                        'created_at': {'$gte': start_date}
                    }},
                    {'$group': {
                        '_id': '$customer_id',
                        'order_count': {'$sum': 1},
                        'total_spent': {'$sum': '$total_amount'}
                    }},
                    {'$sort': {'total_spent': -1}},
                    {'$limit': limit},
                    {'$lookup': {
                        'from': 'customers',
                        'localField': '_id',
                        'foreignField': '_id',
                        'as': 'customer'
                    }},
                    {'$unwind': {'path': '$customer', 'preserveNullAndEmptyArrays': True}}
                ]
                top = list(database['invoices'].aggregate(top_pipeline))
                
                headers = ['Rank', 'Customer Name', 'Email', 'Orders', 'Total Spent']
                for col_num, header in enumerate(headers, 1):
                    cell = ws.cell(row=1, column=col_num, value=header)
                    cell.font = Font(bold=True)
                    cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                    cell.font = Font(bold=True, color="FFFFFF")
                
                for row_num, item in enumerate(top, 2):
                    customer = item.get('customer', {})
                    ws.cell(row=row_num, column=1, value=row_num - 1)
                    ws.cell(row=row_num, column=2, value=customer.get('name', 'Unknown'))
                    ws.cell(row=row_num, column=3, value=customer.get('email', ''))
                    ws.cell(row=row_num, column=4, value=item.get('order_count', 0))
                    ws.cell(row=row_num, column=5, value=float(item.get('total_spent', 0)))
                    
            elif report_type == 'products':
                ws = wb.active
                ws.title = "Top Products"
                
                limit = request.args.get('limit', 50, type=int)
                top_pipeline = [
                    {'$match': {
                        'user_id': user_id_obj,
                        'created_at': {'$gte': start_date}
                    }},
                    {'$unwind': '$items'},
                    {'$group': {
                        '_id': '$items.product_id',
                        'quantity_sold': {'$sum': '$items.quantity'},
                        'revenue': {'$sum': '$items.total'}
                    }},
                    {'$sort': {'quantity_sold': -1}},
                    {'$limit': limit},
                    {'$lookup': {
                        'from': 'products',
                        'localField': '_id',
                        'foreignField': '_id',
                        'as': 'product'
                    }},
                    {'$unwind': {'path': '$product', 'preserveNullAndEmptyArrays': True}}
                ]
                top = list(database['invoices'].aggregate(top_pipeline))
                
                headers = ['Rank', 'Product Name', 'SKU', 'Quantity Sold', 'Revenue']
                for col_num, header in enumerate(headers, 1):
                    cell = ws.cell(row=1, column=col_num, value=header)
                    cell.font = Font(bold=True)
                    cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                    cell.font = Font(bold=True, color="FFFFFF")
                
                for row_num, item in enumerate(top, 2):
                    product = item.get('product', {})
                    ws.cell(row=row_num, column=1, value=row_num - 1)
                    ws.cell(row=row_num, column=2, value=product.get('name', 'Unknown'))
                    ws.cell(row=row_num, column=3, value=product.get('sku', ''))
                    ws.cell(row=row_num, column=4, value=int(item.get('quantity_sold', 0)))
                    ws.cell(row=row_num, column=5, value=float(item.get('revenue', 0)))
            
            # Auto-adjust column widths
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width
            
            output = BytesIO()
            wb.save(output)
            output.seek(0)
            
            return send_file(
                output,
                mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                as_attachment=True,
                download_name=f'report_{report_type}_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
            )
        else:
            # PDF export (if reportlab available)
            if REPORTLAB_AVAILABLE:
                buffer = BytesIO()
                doc = SimpleDocTemplate(buffer, pagesize=A4)
                elements = []
                styles = getSampleStyleSheet()
                
                # Title
                title = Paragraph(f"Sales Report - Last {days} days", styles['Title'])
                elements.append(title)
                elements.append(Spacer(1, 12))
                
                # Get summary data
                revenue_pipeline = [
                    {'$match': {
                        'user_id': user_id_obj,
                        'created_at': {'$gte': start_date}
                    }},
                    {'$group': {'_id': None, 'total': {'$sum': '$total_amount'}}}
                ]
                revenue_result = list(database['invoices'].aggregate(revenue_pipeline))
                total_revenue = revenue_result[0].get('total', 0) if revenue_result else 0.0
                
                total_orders = database['invoices'].count_documents({
                    'user_id': user_id_obj,
                    'created_at': {'$gte': start_date}
                })
                
                # Summary table
                data = [['Metric', 'Value']]
                data.append(['Total Revenue', f'₹{total_revenue:,.2f}'])
                data.append(['Total Orders', str(total_orders)])
                
                table = Table(data)
                table.setStyle(TableStyle([
                    ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                    ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                    ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                    ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                    ('FONTSIZE', (0, 0), (-1, 0), 14),
                    ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                    ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                    ('GRID', (0, 0), (-1, -1), 1, colors.black)
                ]))
                elements.append(table)
                
                doc.build(elements)
                buffer.seek(0)
                
                return send_file(
                    buffer,
                    mimetype='application/pdf',
                    as_attachment=True,
                    download_name=f'report_{report_type}_{datetime.now().strftime("%Y%m%d_%H%M%S")}.pdf'
                )
            else:
                return jsonify({'success': False, 'error': 'PDF export not available'}), 500
                
    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'error': str(e)}), 500
