"""Import/Export routes for CSV and Excel functionality"""
from flask import Blueprint, request, jsonify, send_file
from flask_login import login_required, current_user
from datetime import datetime
import csv
import io
import re
import uuid
from io import StringIO, BytesIO
try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

# Try to import from models, fallback to app_working
try:
    from models import Customer, Product, Order, OrderItem, Invoice, InvoiceItem, StockMovement
    from database import db
    from bson import ObjectId
except ImportError:
    # If using app_working.py, models are defined there
    pass

import_export_bp = Blueprint('import_export', __name__)

# ==================== EXPORT FUNCTIONS ====================

@import_export_bp.route('/export/customers', methods=['GET'])
# @login_required  # TEMPORARILY DISABLED FOR SUBMISSION
def export_customers():
    """Export customers to CSV"""
    try:
        # Get customers (app_working.py doesn't have user_id, so get all active)
        try:
            customers = [Customer.from_dict(doc) for doc in db['customers'].find({'is_active': True})]
        except:
            # Fallback if user_id exists
            if hasattr(current_user, 'id'):
                user_id_obj = ObjectId(current_user.id) if isinstance(current_user.id, str) else current_user.id
                customers = [Customer.from_dict(doc) for doc in db['customers'].find({'user_id': user_id_obj, 'is_active': True})]
            else:
                customers = [Customer.from_dict(doc) for doc in db['customers'].find({'is_active': True})]
        
        # Create CSV in memory
        output = StringIO()
        writer = csv.writer(output)
        
        # Write header
        writer.writerow([
            'Name', 'Email', 'Phone', 'GSTIN', 'Billing Address', 
            'Shipping Address', 'State', 'Pincode', 'Created At', 'Is Active'
        ])
        
        # Write data
        for customer in customers:
            writer.writerow([
                customer.name,
                customer.email,
                customer.phone or '',
                getattr(customer, 'gstin', None) or '',
                getattr(customer, 'billing_address', None) or getattr(customer, 'address', None) or '',
                getattr(customer, 'shipping_address', None) or getattr(customer, 'billing_address', None) or getattr(customer, 'address', None) or '',
                customer.state or '',
                customer.pincode or '',
                customer.created_at.isoformat() if customer.created_at else '',
                'Yes' if customer.is_active else 'No'
            ])
        
        # Create file response
        output.seek(0)
        mem = io.BytesIO()
        mem.write(output.getvalue().encode('utf-8-sig'))  # UTF-8 with BOM for Excel
        mem.seek(0)
        
        return send_file(
            mem,
            mimetype='text/csv',
            as_attachment=True,
            download_name=f'customers_export_{datetime.now().strftime("%Y%m%d_%H%M%S")}.csv'
        )
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500

@import_export_bp.route('/export/products', methods=['GET'])
@login_required
def export_products():
    """Export products to CSV or Excel"""
    try:
        format_type = request.args.get('format', 'excel')  # 'csv' or 'excel'
        
        # Get products
        try:
            user_id_obj = ObjectId(current_user.id) if isinstance(current_user.id, str) else current_user.id
            products = [Product.from_dict(doc) for doc in db['products'].find({'user_id': user_id_obj})]
        except:
            products = [Product.from_dict(doc) for doc in db['products'].find()]
        
        if format_type == 'excel' and OPENPYXL_AVAILABLE:
            # Export to Excel
            wb = Workbook()
            ws = wb.active
            ws.title = "Products"
            
            # Header row with styling
            headers = ['Name', 'SKU', 'HSN Code', 'Description', 'Category', 'Brand', 
                      'Price', 'GST Rate (%)', 'Stock Quantity', 'Min Stock Level', 
                      'Unit', 'Weight', 'Dimensions', 'Image URL', 'Created At']
            
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF")
            
            for col_num, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col_num, value=header)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center', vertical='center')
            
            # Data rows
            for row_num, product in enumerate(products, 2):
                ws.cell(row=row_num, column=1, value=product.name)
                ws.cell(row=row_num, column=2, value=product.sku or '')
                ws.cell(row=row_num, column=3, value=product.hsn_code or '')
                ws.cell(row=row_num, column=4, value=product.description or '')
                ws.cell(row=row_num, column=5, value=product.category or '')
                ws.cell(row=row_num, column=6, value=getattr(product, 'brand', '') or '')
                ws.cell(row=row_num, column=7, value=float(product.price) if product.price else 0)
                ws.cell(row=row_num, column=8, value=float(product.gst_rate) if product.gst_rate else 18)
                ws.cell(row=row_num, column=9, value=int(product.stock_quantity) if product.stock_quantity else 0)
                ws.cell(row=row_num, column=10, value=int(getattr(product, 'min_stock_level', 10)))
                ws.cell(row=row_num, column=11, value=product.unit or 'PCS')
                ws.cell(row=row_num, column=12, value=float(getattr(product, 'weight', 0)) if getattr(product, 'weight', None) else '')
                ws.cell(row=row_num, column=13, value=getattr(product, 'dimensions', '') or '')
                ws.cell(row=row_num, column=14, value=product.image_url or '')
                ws.cell(row=row_num, column=15, value=product.created_at.strftime('%Y-%m-%d %H:%M:%S') if product.created_at else '')
            
            # Auto-adjust column widths
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width
            
            # Save to BytesIO
            output = BytesIO()
            wb.save(output)
            output.seek(0)
            
            return send_file(
                output,
                mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                as_attachment=True,
                download_name=f'products_export_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
            )
        else:
            # Export to CSV (fallback)
            output = StringIO()
            writer = csv.writer(output)
            
            # Write header
            writer.writerow([
                'Name', 'SKU', 'HSN Code', 'Description', 'Category', 'Brand', 'Price', 
                'GST Rate', 'Stock Quantity', 'Min Stock Level', 'Unit', 'Weight', 
                'Dimensions', 'Image URL', 'Created At'
            ])
            
            # Write data
            for product in products:
                writer.writerow([
                    product.name,
                    product.sku or '',
                    product.hsn_code or '',
                    product.description or '',
                    product.category or '',
                    getattr(product, 'brand', '') or '',
                    product.price or 0,
                    product.gst_rate or 18,
                    product.stock_quantity or 0,
                    getattr(product, 'min_stock_level', 10),
                    product.unit or 'PCS',
                    getattr(product, 'weight', '') or '',
                    getattr(product, 'dimensions', '') or '',
                    product.image_url or '',
                    product.created_at.isoformat() if product.created_at else ''
                ])
            
            output.seek(0)
            mem = io.BytesIO()
            mem.write(output.getvalue().encode('utf-8-sig'))
            mem.seek(0)
            
            return send_file(
                mem,
                mimetype='text/csv',
                as_attachment=True,
                download_name=f'products_export_{datetime.now().strftime("%Y%m%d_%H%M%S")}.csv'
            )
    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'error': str(e)}), 500

@import_export_bp.route('/export/orders', methods=['GET'])
# @login_required  # TEMPORARILY DISABLED FOR SUBMISSION
def export_orders():
    """Export orders to CSV"""
    try:
        # Get orders (app_working.py uses admin_id instead of user_id)
        try:
            orders = [Order.from_dict(doc) for doc in db['orders'].find()]
        except:
            if hasattr(current_user, 'id'):
                user_id_obj = ObjectId(current_user.id) if isinstance(current_user.id, str) else current_user.id
                orders = [Order.from_dict(doc) for doc in db['orders'].find({'user_id': user_id_obj})]
            else:
                orders = [Order.from_dict(doc) for doc in db['orders'].find()]
        
        output = StringIO()
        writer = csv.writer(output)
        
        # Write header
        writer.writerow([
            'Order ID', 'Customer Name', 'Customer Email', 'Total Amount', 
            'Status', 'Created At', 'Items (Product:Quantity:Price)'
        ])
        
        # Write data
        for order in orders:
            customer = Customer.find_by_id(order.customer_id)
            order_items = order.items if order.items else []
            items_str = '; '.join([
                f"{Product.find_by_id(item.get('product_id')).name if isinstance(item, dict) and Product.find_by_id(item.get('product_id')) else 'Unknown'}:{item.get('quantity', 0) if isinstance(item, dict) else 0}:{item.get('unit_price', 0) if isinstance(item, dict) else 0}"
                for item in order_items
            ])
            
            writer.writerow([
                order.id,
                customer.name if customer else '',
                customer.email if customer else '',
                order.total_amount or 0,
                order.status or 'pending',
                order.created_at.isoformat() if order.created_at else '',
                items_str
            ])
        
        output.seek(0)
        mem = io.BytesIO()
        mem.write(output.getvalue().encode('utf-8-sig'))
        mem.seek(0)
        
        return send_file(
            mem,
            mimetype='text/csv',
            as_attachment=True,
            download_name=f'orders_export_{datetime.now().strftime("%Y%m%d_%H%M%S")}.csv'
        )
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500

# ==================== IMPORT FUNCTIONS ====================

@import_export_bp.route('/import/customers', methods=['POST'])
# @login_required  # TEMPORARILY DISABLED FOR SUBMISSION
def import_customers():
    """Import customers from CSV"""
    try:
        if 'file' not in request.files:
            return jsonify({'success': False, 'error': 'No file provided'}), 400
        
        file = request.files['file']
        if file.filename == '':
            return jsonify({'success': False, 'error': 'No file selected'}), 400
        
        # Read CSV file
        stream = StringIO(file.stream.read().decode("UTF8"), newline=None)
        csv_reader = csv.DictReader(stream)
        
        imported = 0
        skipped = 0
        errors = []
        
        for row_num, row in enumerate(csv_reader, start=2):  # Start at 2 (row 1 is header)
            try:
                # Required fields
                name = row.get('Name', '').strip()
                email = row.get('Email', '').strip()
                phone = row.get('Phone', '').strip()
                
                if not name or not email:
                    errors.append(f"Row {row_num}: Missing required fields (Name or Email)")
                    skipped += 1
                    continue
                
                # Check if customer already exists
                existing = Customer.find_by_email(email)
                if existing:
                    errors.append(f"Row {row_num}: Customer with email {email} already exists")
                    skipped += 1
                    continue
                
                # Create customer (handle both model structures)
                address = row.get('Billing Address', '').strip() or row.get('Address', '').strip() or ''
                customer = Customer(
                    name=name,
                    email=email,
                    phone=phone or '',
                    address=address,
                    state=row.get('State', '').strip() or '',
                    pincode=row.get('Pincode', '').strip() or '',
                    is_active=row.get('Is Active', 'Yes').strip().lower() in ['yes', 'true', '1', 'y']
                )
                
                # Set additional fields if they exist in the model
                if hasattr(customer, 'gstin'):
                    customer.gstin = row.get('GSTIN', '').strip() or None
                if hasattr(customer, 'billing_address'):
                    customer.billing_address = address
                if hasattr(customer, 'shipping_address'):
                    customer.shipping_address = row.get('Shipping Address', '').strip() or None
                if hasattr(customer, 'user_id'):
                    customer.user_id = current_user.id if hasattr(current_user, 'id') else None
                
                # Set default password
                default_password = row.get('Password', 'default123').strip() or 'default123'
                customer.set_password(default_password)
                
                customer.save()
                imported += 1
                
            except Exception as e:
                errors.append(f"Row {row_num}: {str(e)}")
                skipped += 1
        
        return jsonify({
            'success': True,
            'imported': imported,
            'skipped': skipped,
            'errors': errors[:10]  # Limit errors to first 10
        })
        
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500

@import_export_bp.route('/import/products', methods=['POST'])
@login_required
def import_products():
    """Import products from CSV or Excel"""
    try:
        if 'file' not in request.files:
            return jsonify({'success': False, 'error': 'No file provided'}), 400
        
        file = request.files['file']
        if file.filename == '':
            return jsonify({'success': False, 'error': 'No file selected'}), 400
        
        # Check file extension
        filename = file.filename.lower()
        is_excel = filename.endswith('.xlsx') or filename.endswith('.xls')
        
        rows = []
        
        file_bytes = file.read()
        if not file_bytes:
            return jsonify({'success': False, 'error': 'Uploaded file is empty'}), 400
        
        if is_excel and OPENPYXL_AVAILABLE:
            # Read Excel file
            from openpyxl import load_workbook
            wb = load_workbook(BytesIO(file_bytes), read_only=True)
            ws = wb.active
            
            # Get headers from first row - handle None cells
            headers = []
            try:
                for cell in ws[1]:
                    if cell and cell.value is not None:
                        headers.append(str(cell.value).strip())
            except Exception as e:
                return jsonify({'success': False, 'error': f'Error reading headers: {str(e)}'}), 400
            
            if not headers:
                return jsonify({'success': False, 'error': 'No headers found in Excel file'}), 400
            
            # Read data rows
            for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                try:
                    # Skip None rows or rows with all None values
                    if row is None:
                        continue
                    if not any(cell is not None and str(cell).strip() for cell in row):
                        continue
                    
                    row_dict = {}
                    for idx, header in enumerate(headers):
                        if header:  # Only process non-empty headers
                            if idx < len(row) and row[idx] is not None:
                                row_dict[header] = str(row[idx]).strip()
                            else:
                                row_dict[header] = ''
                    if row_dict:  # Only add non-empty row dicts
                        rows.append(row_dict)
                except Exception as e:
                    # Skip problematic rows but log the error
                    print(f"Warning: Error processing row {row_num}: {e}")
                    continue
        else:
            # Read CSV file
            stream = StringIO(file_bytes.decode("UTF8"), newline=None)
            csv_reader = csv.DictReader(stream)
            rows = list(csv_reader)
        
        imported = 0
        skipped = 0
        errors = []
        
        for row_num, row in enumerate(rows, start=2):
            try:
                # Validate row is not None and is a dict
                if row is None:
                    errors.append(f"Row {row_num}: Row is None")
                    skipped += 1
                    continue
                if not isinstance(row, dict):
                    errors.append(f"Row {row_num}: Invalid row format")
                    skipped += 1
                    continue
                
                # Handle vegetable-specific columns - check for various possible column name variations
                # Vegetable Name (English) - could be "Vegetable Name", "Product Name", "Name", "ple Name", etc.
                vegetable_name = ''
                # Ensure row has keys() method (is a dict)
                row_keys = row.keys() if hasattr(row, 'keys') and isinstance(row, dict) else []
                for key in row_keys:
                    key_lower = key.lower().strip()
                    # Check for vegetable name columns (excluding Hindi)
                    if ('vegetable' in key_lower or 'product' in key_lower or 'ple name' in key_lower) and 'name' in key_lower and 'hindi' not in key_lower:
                        val = str(row.get(key, '') or '').strip()
                        if val:
                            vegetable_name = val
                            break
                    # Also check for just "Name" if it's the first column or doesn't contain hindi
                    elif key_lower == 'name' or (key_lower.endswith('name') and 'hindi' not in key_lower):
                        val = str(row.get(key, '') or '').strip()
                        if val:
                            vegetable_name = val
                            # Don't break here, continue checking for more specific matches
                
                # Fallback to 'Name' if vegetable name not found
                if not vegetable_name:
                    name = str(row.get('Name', '') or '').strip()
                    vegetable_name = name
                    # If still not found, try first column
                    if not vegetable_name and row:
                        first_key = list(row.keys())[0] if row.keys() else None
                        if first_key:
                            vegetable_name = str(row.get(first_key, '') or '').strip()
                
                # Vegetable Name (Hindi)
                vegetable_name_hindi = ''
                for key in row_keys:
                    key_lower = key.lower().strip()
                    if 'hindi' in key_lower or '(hindi)' in key_lower or 'ble name' in key_lower:
                        val = str(row.get(key, '') or '').strip()
                        if val:
                            vegetable_name_hindi = val
                            break
                
                if not vegetable_name:
                    errors.append(f"Row {row_num}: Missing required field (Vegetable Name/Name)")
                    skipped += 1
                    continue
                
                # Generate SKU if not provided
                sku = str(row.get('SKU', '') or '').strip()
                if not sku:
                    sku = f"SKU-{vegetable_name[:10].upper().replace(' ', '-')}-{datetime.now().strftime('%Y%m%d%H%M%S')[-8:]}"
                
                # Check if product with same SKU exists
                try:
                    user_id_obj = ObjectId(current_user.id) if isinstance(current_user.id, str) else current_user.id
                    existing_doc = db['products'].find_one({'sku': sku, 'user_id': user_id_obj})
                    if existing_doc:
                        errors.append(f"Row {row_num}: Product with SKU {sku} already exists")
                        skipped += 1
                        continue
                except Exception as e:
                    # If check fails, continue with import
                    pass
                
                # Parse vegetable-specific quantity and rate fields
                quantity_gm = None
                quantity_kg = None
                rate_per_gm = None
                rate_per_kg = None
                
                # Quantity (gm) - handle truncated headers like "uantity (gm"
                for key in row_keys:
                    key_lower = key.lower().strip()
                    if 'quantity' in key_lower and ('gm' in key_lower or 'g)' in key_lower) and 'kg' not in key_lower:
                        try:
                            qty_val = str(row.get(key, '') or '').strip()
                            quantity_gm = float(qty_val) if qty_val else None
                        except (ValueError, TypeError):
                            pass
                        if quantity_gm is not None:
                            break
                
                # Quantity (kg) - handle truncated headers like "uantity (k"
                for key in row_keys:
                    key_lower = key.lower().strip()
                    if 'quantity' in key_lower and ('kg' in key_lower or 'k)' in key_lower):
                        try:
                            qty_val = str(row.get(key, '') or '').strip()
                            quantity_kg = float(qty_val) if qty_val else None
                        except (ValueError, TypeError):
                            pass
                        if quantity_kg is not None:
                            break
                
                # Rate (per gm) - handle truncated headers like "ate (per gnate"
                for key in row_keys:
                    key_lower = key.lower().strip()
                    if 'rate' in key_lower and ('per' in key_lower or 'per' in key) and ('gm' in key_lower or 'gnate' in key_lower or 'g)' in key_lower) and 'kg' not in key_lower:
                        try:
                            rate_val = str(row.get(key, '') or '').strip()
                            rate_per_gm = float(rate_val) if rate_val else None
                        except (ValueError, TypeError):
                            pass
                        if rate_per_gm is not None:
                            break
                
                # Rate (per kg) - handle truncated headers like "(per kg)"
                for key in row_keys:
                    key_lower = key.lower().strip()
                    if ('rate' in key_lower or key_lower.startswith('(')) and ('per' in key_lower or 'per' in key) and ('kg' in key_lower or 'k)' in key_lower):
                        try:
                            rate_val = str(row.get(key, '') or '').strip()
                            rate_per_kg = float(rate_val) if rate_val else None
                        except (ValueError, TypeError):
                            pass
                        if rate_per_kg is not None:
                            break
                
                # Calculate price from rate_per_kg if available, otherwise use rate_per_gm * 1000
                # If neither is available, try to get Price directly
                price = 0
                if rate_per_kg:
                    price = rate_per_kg
                elif rate_per_gm:
                    price = rate_per_gm * 1000  # Convert per gm to per kg
                else:
                    # Fallback to Price column
                    try:
                        price_val = str(row.get('Price', '') or row.get('Price', 0) or '0').strip()
                        price = float(price_val) if price_val else 0
                    except (ValueError, TypeError):
                        price = 0
                
                # Calculate stock_quantity from quantity_kg if available, otherwise use quantity_gm / 1000
                stock_quantity = 0
                if quantity_kg:
                    stock_quantity = int(quantity_kg * 1000)  # Convert kg to grams for stock
                elif quantity_gm:
                    stock_quantity = int(quantity_gm)
                else:
                    # Fallback to Stock Quantity column
                    try:
                        stock_val = str(row.get('Stock Quantity', '') or '0').strip()
                        stock_quantity = int(stock_val) if stock_val else 0
                    except (ValueError, TypeError):
                        stock_quantity = 0
                
                # Parse GST rate
                try:
                    gst_val = str(row.get('GST Rate', '') or row.get('GST Rate (%)', '') or '18').strip()
                    gst_rate = float(gst_val) if gst_val else 18
                except (ValueError, TypeError):
                    gst_rate = 18
                
                # Create product (handle both model structures)
                # Get user_id for admin_id compatibility
                user_id_for_product = current_user.id if hasattr(current_user, 'id') else None
                if not user_id_for_product:
                    return jsonify({'success': False, 'error': 'User not authenticated'}), 401
                
                product = Product(
                    user_id=user_id_for_product,
                    admin_id=user_id_for_product,  # Set admin_id to same as user_id for backward compatibility
                    name=vegetable_name,  # Use vegetable_name as the main name
                    price=price,
                    gst_rate=gst_rate,
                    stock_quantity=stock_quantity,
                    sku=sku
                )
                
                # Set vegetable-specific fields (with error handling)
                try:
                    if hasattr(product, 'vegetable_name'):
                        product.vegetable_name = vegetable_name
                except Exception:
                    pass  # Column might not exist, ignore
                
                try:
                    if hasattr(product, 'vegetable_name_hindi'):
                        product.vegetable_name_hindi = vegetable_name_hindi if vegetable_name_hindi else None
                except Exception:
                    pass
                
                try:
                    if hasattr(product, 'quantity_gm'):
                        product.quantity_gm = quantity_gm
                except Exception:
                    pass
                
                try:
                    if hasattr(product, 'quantity_kg'):
                        product.quantity_kg = quantity_kg
                except Exception:
                    pass
                
                try:
                    if hasattr(product, 'rate_per_gm'):
                        product.rate_per_gm = rate_per_gm
                except Exception:
                    pass
                
                try:
                    if hasattr(product, 'rate_per_kg'):
                        product.rate_per_kg = rate_per_kg
                except Exception:
                    pass
                
                # Set additional fields if they exist
                if hasattr(product, 'hsn_code'):
                    hsn_val = str(row.get('HSN Code', '') or '').strip()
                    product.hsn_code = hsn_val if hsn_val else None
                if hasattr(product, 'description'):
                    desc_val = str(row.get('Description', '') or '').strip()
                    product.description = desc_val if desc_val else None
                if hasattr(product, 'category'):
                    cat_val = str(row.get('Category', '') or '').strip()
                    product.category = cat_val if cat_val else None
                if hasattr(product, 'brand'):
                    brand_val = str(row.get('Brand', '') or '').strip()
                    product.brand = brand_val if brand_val else None
                if hasattr(product, 'unit'):
                    unit_val = str(row.get('Unit', '') or '').strip()
                    product.unit = unit_val if unit_val else 'KG'  # Default to KG for vegetables
                if hasattr(product, 'image_url'):
                    img_val = str(row.get('Image URL', '') or '').strip()
                    product.image_url = img_val if img_val else None
                if hasattr(product, 'min_stock_level'):
                    try:
                        min_stock_val = str(row.get('Min Stock Level', '') or '').strip()
                        product.min_stock_level = int(min_stock_val) if min_stock_val else 10
                    except (ValueError, TypeError):
                        product.min_stock_level = 10
                if hasattr(product, 'weight'):
                    try:
                        weight_val = str(row.get('Weight', '') or '').strip()
                        product.weight = float(weight_val) if weight_val else None
                    except (ValueError, TypeError):
                        product.weight = None
                if hasattr(product, 'dimensions'):
                    dim_val = str(row.get('Dimensions', '') or '').strip()
                    product.dimensions = dim_val if dim_val else None
                # user_id and admin_id are already set during Product creation above
                
                product.save()
                imported += 1
                
            except Exception as e:
                errors.append(f"Row {row_num}: {str(e)}")
                skipped += 1
        
        return jsonify({
            'success': True,
            'imported': imported,
            'skipped': skipped,
            'errors': errors[:10]
        })
        
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500

@import_export_bp.route('/import/orders', methods=['POST'])
# @login_required  # TEMPORARILY DISABLED FOR SUBMISSION
def import_orders():
    """Import orders from CSV"""
    try:
        if 'file' not in request.files:
            return jsonify({'success': False, 'error': 'No file provided'}), 400
        
        file = request.files['file']
        if file.filename == '':
            return jsonify({'success': False, 'error': 'No file selected'}), 400
        
        stream = StringIO(file.stream.read().decode("UTF8"), newline=None)
        csv_reader = csv.DictReader(stream)
        
        imported = 0
        skipped = 0
        errors = []
        
        for row_num, row in enumerate(csv_reader, start=2):
            try:
                customer_email = row.get('Customer Email', '').strip()
                if not customer_email:
                    errors.append(f"Row {row_num}: Missing Customer Email")
                    skipped += 1
                    continue
                
                # Find customer
                try:
                    customer = Customer.find_by_email(customer_email)
                except:
                    if hasattr(current_user, 'id'):
                        user_id_obj = ObjectId(current_user.id) if isinstance(current_user.id, str) else current_user.id
                        customer_doc = db['customers'].find_one({'email': customer_email, 'user_id': user_id_obj})
                        customer = Customer.from_dict(customer_doc) if customer_doc else None
                    else:
                        customer = Customer.find_by_email(customer_email)
                if not customer:
                    errors.append(f"Row {row_num}: Customer with email {customer_email} not found")
                    skipped += 1
                    continue
                
                # Parse items
                items_str = row.get('Items (Product:Quantity:Price)', '').strip()
                if not items_str:
                    errors.append(f"Row {row_num}: Missing items")
                    skipped += 1
                    continue
                
                # Calculate total
                total_amount = 0
                items_list = []
                
                for item_str in items_str.split(';'):
                    parts = item_str.strip().split(':')
                    if len(parts) >= 2:
                        product_name = parts[0].strip()
                        quantity = int(parts[1].strip()) if len(parts) > 1 else 1
                        price = float(parts[2].strip()) if len(parts) > 2 else 0
                        
                        # Find product
                        try:
                            product_doc = db['products'].find_one({'name': product_name})
                            product = Product.from_dict(product_doc) if product_doc else None
                        except:
                            if hasattr(current_user, 'id'):
                                user_id_obj = ObjectId(current_user.id) if isinstance(current_user.id, str) else current_user.id
                                product_doc = db['products'].find_one({'name': product_name, 'user_id': user_id_obj})
                                product = Product.from_dict(product_doc) if product_doc else None
                            else:
                                product = None
                        if product:
                            items_list.append({
                                'product': product,
                                'quantity': quantity,
                                'price': price or product.price
                            })
                            total_amount += (price or product.price) * quantity
                
                if not items_list:
                    errors.append(f"Row {row_num}: No valid items found")
                    skipped += 1
                    continue
                
                # Create order (handle both model structures)
                order = Order(
                    customer_id=customer.id,
                    total_amount=total_amount,
                    status=row.get('Status', 'pending').strip() or 'pending'
                )
                
                # Set admin_id or user_id based on model
                if hasattr(order, 'admin_id'):
                    order.admin_id = current_user.id if hasattr(current_user, 'id') else None
                if hasattr(order, 'user_id'):
                    order.user_id = current_user.id if hasattr(current_user, 'id') else None
                order.save()
                
                # Create order items
                order_items_list = []
                for item_data in items_list:
                    order_item = OrderItem(
                        order_id=order.id,
                        product_id=item_data['product'].id,
                        quantity=item_data['quantity'],
                        unit_price=item_data['price'],
                        total=item_data['quantity'] * item_data['price']
                    )
                    order_item.save()
                    order_items_list.append(order_item.to_dict())
                
                # Update order with items
                order.items = order_items_list
                order.save()
                
                imported += 1
                
            except Exception as e:
                errors.append(f"Row {row_num}: {str(e)}")
                skipped += 1
        
        return jsonify({
            'success': True,
            'imported': imported,
            'skipped': skipped,
            'errors': errors[:10]
        })
        
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500

@import_export_bp.route('/import/stock', methods=['POST'])
@login_required
def import_stock():
    """Import stock movements from CSV or Excel for bulk stock in/out"""
    try:
        if 'file' not in request.files:
            return jsonify({'success': False, 'error': 'No file provided'}), 400
        
        file = request.files['file']
        if file.filename == '':
            return jsonify({'success': False, 'error': 'No file selected'}), 400
        
        movement_type = request.form.get('movement_type', 'in')
        if movement_type not in ['in', 'out']:
            movement_type = 'in'
        
        filename = file.filename.lower()
        is_excel = filename.endswith('.xlsx') or filename.endswith('.xls')
        rows = []
        
        file_bytes = file.read()
        if not file_bytes:
            return jsonify({'success': False, 'error': 'Uploaded file is empty'}), 400
        
        if is_excel and OPENPYXL_AVAILABLE:
            from openpyxl import load_workbook
            wb = load_workbook(BytesIO(file_bytes), read_only=True)
            ws = wb.active
            
            # Get headers - handle None cells
            headers = []
            try:
                header_row = ws[1]
                if header_row is None:
                    return jsonify({'success': False, 'error': 'Header row is None'}), 400
                for cell in header_row:
                    if cell is not None and hasattr(cell, 'value') and cell.value is not None:
                        headers.append(str(cell.value).strip())
            except Exception as e:
                import traceback
                traceback.print_exc()
                return jsonify({'success': False, 'error': f'Error reading headers: {str(e)}'}), 400
            
            if not headers:
                return jsonify({'success': False, 'error': 'No headers found in Excel file'}), 400
            
            # Read data rows
            for row_num, row_tuple in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                try:
                    # Skip None rows
                    if row_tuple is None:
                        print(f"Warning: Row {row_num} tuple is None, skipping")
                        continue
                    
                    # Ensure row is iterable (tuple or list)
                    if not isinstance(row_tuple, (tuple, list)):
                        print(f"Warning: Row {row_num} is not iterable (type: {type(row_tuple)}), skipping")
                        continue
                    
                    # Check if row has any non-empty values
                    has_data = False
                    try:
                        for cell in row_tuple:
                            if cell is not None:
                                try:
                                    cell_str = str(cell).strip()
                                    if cell_str:
                                        has_data = True
                                        break
                                except:
                                    pass
                    except Exception as e:
                        print(f"Warning: Error checking row {row_num} data: {e}")
                        continue
                    
                    if not has_data:
                        continue
                    
                    # Build row dictionary safely
                    row_dict = {}
                    try:
                        for idx, header in enumerate(headers):
                            if header and isinstance(header, str):  # Only process valid string headers
                                try:
                                    if idx < len(row_tuple):
                                        cell_value = row_tuple[idx]
                                        if cell_value is not None:
                                            row_dict[str(header)] = str(cell_value).strip()
                                        else:
                                            row_dict[str(header)] = ''
                                    else:
                                        row_dict[str(header)] = ''
                                except (IndexError, TypeError) as e:
                                    row_dict[str(header)] = ''
                    except Exception as e:
                        print(f"Warning: Error building row_dict for row {row_num}: {e}")
                        import traceback
                        traceback.print_exc()
                        continue
                    
                    # Only add valid non-empty row dicts
                    if row_dict and isinstance(row_dict, dict) and len(row_dict) > 0:
                        rows.append(row_dict)
                    else:
                        print(f"Warning: Row {row_num} produced invalid or empty row_dict, skipping")
                except Exception as e:
                    # Skip problematic rows but log the error
                    import traceback
                    print(f"Warning: Error processing row {row_num}: {e}")
                    traceback.print_exc()
                    continue
        else:
            # Read CSV file
            stream = StringIO(file_bytes.decode("UTF8"), newline=None)
            csv_reader = csv.DictReader(stream)
            rows = []
            for row in csv_reader:
                # Validate CSV rows are not None
                if row is not None and isinstance(row, dict):
                    rows.append(row)
        
        if not rows:
            return jsonify({'success': False, 'error': 'No data found in file'}), 400
        
        # Get database connection dynamically (once, outside the loop)
        try:
            from models import get_db
            database = get_db()
            if database is None:
                return jsonify({'success': False, 'error': 'Database not initialized'}), 500
        except Exception as e:
            import traceback
            traceback.print_exc()
            return jsonify({'success': False, 'error': f'Database error: {str(e)}'}), 500
        
        imported = 0
        skipped = 0
        errors = []
        
        def get_field_value(row_dict, include_terms, exclude_terms=None):
            if row_dict is None:
                return ''
            if not isinstance(row_dict, dict):
                return ''
            exclude_terms = exclude_terms or []
            try:
                # Safely get items from dict
                items = row_dict.items() if hasattr(row_dict, 'items') else []
                for key, value in items:
                    if key is None:
                        continue
                    try:
                        key_str = str(key).lower().strip()
                        if not key_str:
                            continue
                        if all(term in key_str for term in include_terms):
                            if any(term in key_str for term in exclude_terms):
                                continue
                            return str(value or '').strip()
                    except Exception as e:
                        print(f"Error processing key in get_field_value: {e}")
                        continue
            except (TypeError, AttributeError) as e:
                print(f"Error in get_field_value (TypeError/AttributeError): {e}")
                import traceback
                traceback.print_exc()
            except Exception as e:
                print(f"Error in get_field_value: {e}")
                import traceback
                traceback.print_exc()
            return ''

        def parse_float(val):
            if val is None or val == '':
                return None
            try:
                if isinstance(val, str):
                    cleaned = val.replace(',', '').strip()
                    if cleaned == '':
                        return None
                    return float(cleaned)
                return float(val)
            except (ValueError, TypeError):
                return None

        for row_num, row in enumerate(rows, start=2):
            try:
                # Validate row is not None and is a dict
                if row is None:
                    errors.append(f"Row {row_num}: Row is None")
                    skipped += 1
                    continue
                if not isinstance(row, dict):
                    errors.append(f"Row {row_num}: Invalid row format (expected dict, got {type(row).__name__})")
                    skipped += 1
                    continue
                
                # Ensure row has keys() method and is accessible
                try:
                    _ = row.keys()
                except (AttributeError, TypeError) as e:
                    errors.append(f"Row {row_num}: Cannot access row keys: {str(e)}")
                    skipped += 1
                    continue
                
                product = None
                # Safely get values with None handling
                product_id = str(row.get('Product ID') or '').strip()
                sku = str(row.get('SKU') or '').strip()
                product_name = str(row.get('Product Name') or '').strip() or str(row.get('Name') or '').strip()
                vegetable_name = get_field_value(row, ['vegetable', 'name'], ['hindi']) or product_name
                vegetable_name_hindi = get_field_value(row, ['vegetable', 'name', 'hindi'])
                quantity_gm_str = get_field_value(row, ['quantity', 'gm'])
                quantity_kg_str = get_field_value(row, ['quantity', 'kg'])
                rate_per_gm_str = get_field_value(row, ['rate', 'gm'])
                rate_per_kg_str = get_field_value(row, ['rate', 'kg'])
                
                user_id = current_user.id if hasattr(current_user, 'id') else None
                if not user_id:
                    errors.append(f"Row {row_num}: User not authenticated")
                    skipped += 1
                    continue
                
                user_id_obj = ObjectId(user_id) if isinstance(user_id, str) else user_id
                
                if product_id:
                    try:
                        product_doc = database['products'].find_one({'_id': ObjectId(int(product_id)), 'user_id': user_id_obj, 'is_active': True})
                        product = Product.from_dict(product_doc) if product_doc else None
                    except (ValueError, TypeError):
                        pass
                
                if not product and sku:
                    product_doc = database['products'].find_one({'sku': sku, 'user_id': user_id_obj, 'is_active': True})
                    product = Product.from_dict(product_doc) if product_doc else None
                
                if not product and product_name:
                    product_doc = database['products'].find_one({'name': product_name, 'user_id': user_id_obj, 'is_active': True})
                    product = Product.from_dict(product_doc) if product_doc else None

                if not product and vegetable_name:
                    product_doc = database['products'].find_one({'name': vegetable_name, 'user_id': user_id_obj, 'is_active': True})
                    product = Product.from_dict(product_doc) if product_doc else None

                price_per_kg = parse_float(rate_per_kg_str)
                price_per_gm = parse_float(rate_per_gm_str)
                if price_per_kg is None and price_per_gm is not None:
                    price_per_kg = price_per_gm * 1000
                
                if not product and vegetable_name:
                    base_name = vegetable_name.strip()
                    sku_slug = re.sub(r'[^A-Z0-9]', '', base_name.upper())[:10] or 'VEG'
                    unique_sku = f"{sku_slug}-{datetime.utcnow().strftime('%H%M%S')}"
                    counter = 1
                    gen_sku = unique_sku
                    while database['products'].find_one({'sku': gen_sku, 'user_id': user_id_obj}):
                        gen_sku = f"{unique_sku}-{counter}"
                        counter += 1
                    
                    product = Product(
                        user_id=user_id,
                        admin_id=user_id,  # Set admin_id to same as user_id for backward compatibility
                        name=base_name,
                        sku=gen_sku,
                        category='Vegetables',
                        unit='GMS',
                        price=price_per_kg or 0.0,
                        gst_rate=0.0,
                        stock_quantity=0,
                        min_stock_level=10,
                        is_active=True
                    )
                    product.save()
                
                if not product:
                    errors.append(f"Row {row_num}: Product not found")
                    skipped += 1
                    continue
                
                quantity_str = str(row.get('Quantity') or '').strip() or str(row.get('Qty') or '').strip()
                if not quantity_str:
                    quantity_str = quantity_gm_str or ''
                
                quantity = None
                if quantity_str:
                    quantity = parse_float(quantity_str)
                if quantity is None and quantity_kg_str:
                    kg_val = parse_float(quantity_kg_str)
                    if kg_val is not None:
                        quantity = kg_val * 1000
                if quantity is None and quantity_gm_str:
                    quantity = parse_float(quantity_gm_str)
                
                if quantity is None:
                    errors.append(f"Row {row_num}: Quantity is required")
                    skipped += 1
                    continue
                
                try:
                    quantity_int = int(round(float(quantity)))
                    if quantity_int <= 0:
                        errors.append(f"Row {row_num}: Quantity must be greater than 0")
                        skipped += 1
                        continue
                except (ValueError, TypeError):
                    errors.append(f"Row {row_num}: Invalid quantity")
                    skipped += 1
                    continue
                
                if vegetable_name:
                    product.vegetable_name = vegetable_name
                if vegetable_name_hindi:
                    product.vegetable_name_hindi = vegetable_name_hindi
                
                quantity_gm_val = parse_float(quantity_gm_str)
                quantity_kg_val = parse_float(quantity_kg_str)
                if quantity_gm_val is not None:
                    product.quantity_gm = quantity_gm_val
                if quantity_kg_val is not None:
                    product.quantity_kg = quantity_kg_val
                rate_per_gm_val = parse_float(rate_per_gm_str)
                rate_per_kg_val = parse_float(rate_per_kg_str)
                if rate_per_gm_val is not None:
                    product.rate_per_gm = rate_per_gm_val
                if rate_per_kg_val is not None:
                    product.rate_per_kg = rate_per_kg_val
                if price_per_kg is not None:
                    product.price = price_per_kg
                
                if movement_type == 'out' and product.stock_quantity < quantity_int:
                    errors.append(f"Row {row_num}: Insufficient stock for {product.name}")
                    skipped += 1
                    continue
                
                reference = str(row.get('Reference') or '').strip() or f'Bulk import - Row {row_num}'
                notes = str(row.get('Notes') or '').strip() or f'Stock {movement_type} from Excel import'
                
                movement = StockMovement(
                    product_id=product.id,
                    movement_type=movement_type,
                    quantity=quantity_int,
                    reference=reference,
                    notes=notes
                )
                
                if movement_type == 'in':
                    product.stock_quantity += quantity_int
                else:
                    product.stock_quantity -= quantity_int
                
                product.updated_at = datetime.utcnow()
                product.save()
                
                movement.save()
                imported += 1
                
            except TypeError as e:
                # Handle 'NoneType' object is not subscriptable errors
                if "'NoneType' object is not subscriptable" in str(e) or "not subscriptable" in str(e):
                    errors.append(f"Row {row_num}: Invalid data format - row contains None values")
                else:
                    errors.append(f"Row {row_num}: {str(e)}")
                skipped += 1
                import traceback
                print(f"Error in row {row_num}: {e}")
                traceback.print_exc()
                continue
            except Exception as e:
                error_msg = str(e)
                if "'NoneType' object is not subscriptable" in error_msg or "not subscriptable" in error_msg:
                    errors.append(f"Row {row_num}: Invalid data format - row contains None values")
                else:
                    errors.append(f"Row {row_num}: {error_msg}")
                skipped += 1
                import traceback
                print(f"Error in row {row_num}: {e}")
                traceback.print_exc()
                continue
        
        return jsonify({
            'success': True,
            'imported': imported,
            'skipped': skipped,
            'errors': errors[:20]
        })
        
    except Exception as e:
        import traceback
        print(f"Error in stock import: {str(e)}")
        traceback.print_exc()
        return jsonify({'success': False, 'error': str(e)}), 500


